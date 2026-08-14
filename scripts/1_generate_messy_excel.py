import random
from datetime import date, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

random.seed(7)

regions = ["North", "South", "East", "West"]
categories_clean = ["Electronics", "Apparel", "Home & Kitchen", "Sports", "Beauty"]
# messy variants that all mean the same category — the cleaning script must fix these
category_variants = {
    "Electronics": ["Electronics", "electronics", "ELECTRONICS", "Electronics ", " Electronics"],
    "Apparel": ["Apparel", "apparel", "APPAREL", "Appare l", "Apparel "],
    "Home & Kitchen": ["Home & Kitchen", "home & kitchen", "Home and Kitchen", "HOME & KITCHEN"],
    "Sports": ["Sports", "sports", "SPORTS", "Sports "],
    "Beauty": ["Beauty", "beauty", "BEAUTY", " Beauty"],
}
channels = ["Online", "In-Store", "online", "IN-STORE", "In Store"]
price_ranges = {
    "Electronics": (80, 600), "Apparel": (15, 120), "Home & Kitchen": (20, 250),
    "Sports": (10, 200), "Beauty": (8, 90),
}

date_formats = ["%Y-%m-%d", "%m/%d/%Y", "%d-%b-%Y", "%m/%d/%y"]

customers = [f"CUST-{1000+i}" for i in range(120)]
start = date(2025, 1, 1)
end = date(2025, 12, 31)
n_days = (end - start).days

rows = []
order_id = 700001
for d_off in range(n_days + 1):
    current = start + timedelta(days=d_off)
    n_orders = max(0, int(random.gauss(6, 2)))
    for _ in range(n_orders):
        cat_clean = random.choice(categories_clean)
        cat_display = random.choice(category_variants[cat_clean])
        lo, hi = price_ranges[cat_clean]
        unit_price = round(random.uniform(lo, hi), 2)
        units = random.choice([1, 1, 1, 2, 2, 3])
        revenue = round(unit_price * units, 2)
        cost = round(revenue * random.uniform(0.55, 0.75), 2)

        # pick a random date string format to simulate inconsistent exports
        date_str = current.strftime(random.choice(date_formats))

        region = random.choice(regions)
        channel = random.choice(channels)
        customer = random.choice(customers)

        rows.append({
            "OrderID": order_id,
            "OrderDate": date_str,
            "CustomerID": customer,
            "Region": region,
            "Category": cat_display,
            "Channel": channel,
            "Units": units,
            "UnitPrice": unit_price,
            "Revenue": revenue,
            "Cost": cost,
        })
        order_id += 1

# ---- Inject data quality problems ----
# 1. Duplicate ~2% of rows exactly (common export glitch)
dupes = random.sample(rows, k=int(len(rows) * 0.02))
rows.extend(dupes)

# 2. Blank out ~1.5% of Region values
for r in random.sample(rows, k=int(len(rows) * 0.015)):
    r["Region"] = None

# 3. Blank out ~1% of CustomerID
for r in random.sample(rows, k=int(len(rows) * 0.01)):
    r["CustomerID"] = ""

# 4. A few negative/zero unit prices (entry errors) ~0.5%
for r in random.sample(rows, k=int(len(rows) * 0.005)):
    r["UnitPrice"] = -abs(r["UnitPrice"])
    r["Revenue"] = round(r["UnitPrice"] * r["Units"], 2)

# 5. A few rows with Units = 0 (should be excluded)
for r in random.sample(rows, k=int(len(rows) * 0.005)):
    r["Units"] = 0
    r["Revenue"] = 0.0

# 6. Extra whitespace injected into CustomerID and Region for some rows
for r in random.sample(rows, k=int(len(rows) * 0.02)):
    if r["CustomerID"]:
        r["CustomerID"] = f"  {r['CustomerID']} "
    if r["Region"]:
        r["Region"] = f" {r['Region']}"

random.shuffle(rows)

# ---- Write to Excel ----
wb = Workbook()
ws = wb.active
ws.title = "Sales Export"
headers = list(rows[0].keys())
ws.append(headers)
for c in range(1, len(headers) + 1):
    cell = ws.cell(row=1, column=c)
    cell.font = Font(name="Arial", bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="404040")

for r in rows:
    ws.append(list(r.values()))

for i, w in enumerate([10, 12, 14, 10, 18, 12, 8, 11, 11, 11], start=1):
    ws.column_dimensions[chr(64 + i)].width = w

wb.save("raw/raw_sales_export.xlsx")
print(f"Wrote {len(rows)} rows (including injected duplicates/errors) to raw/raw_sales_export.xlsx")
